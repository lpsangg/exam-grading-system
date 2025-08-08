from fastapi import APIRouter, UploadFile, File, Request
from fastapi.responses import JSONResponse, FileResponse
import os
import pandas as pd
from typing import List
import logging
import json

# Import image processing functions
from utils.image_processing import image_processing
from utils.detectCodeBox import detect_code_box
from utils.detectInfo import detect_name_student, detect_id_student, detect_index_student
from utils.detectGrade import predict_grade
from utils.automatic_exam_grading import calculate_score
from utils.student_validation import validate_and_correct_student_info
from utils.processing_result_file import process_df_student

def normalize_path(path):
    """Normalize path separators to forward slashes for web compatibility"""
    return path.replace("\\", "/")

# Cấu hình logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter()

# Upload đáp án (Excel)
@router.post('/api/upload_answer_key')
async def upload_answer_key(file: UploadFile = File(...)):
    logger.info(f"Received file upload request: {file.filename}")
    
    if not file:
        logger.error("No file provided")
        return JSONResponse({'error': 'No file part'}, status_code=400)
    
    # Kiểm tra định dạng file
    if not file.filename.endswith('.xlsx'):
        logger.error(f"Invalid file format: {file.filename}")
        return JSONResponse({'error': 'Chỉ chấp nhận file Excel (.xlsx)'}, status_code=400)
    
    try:
        filename = file.filename
        save_path = os.path.join('uploads', 'key', filename)
        
        logger.info(f"Saving file to: {save_path}")
        
        # Đọc file content
        content = await file.read()
        logger.info(f"File size: {len(content)} bytes")
        
        # Đảm bảo thư mục tồn tại
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        
        # Lưu file
        with open(save_path, 'wb') as f:
            f.write(content)
        
        logger.info("File saved successfully")
        
        # Đọc file Excel để lấy exam codes và answer keys
        import io
        try:
            df = pd.read_excel(io.BytesIO(content))
            logger.info(f"Excel file read successfully. Shape: {df.shape}")
            logger.info(f"Columns: {list(df.columns)}")
            
            # Lấy mã đề và đáp án
            exam_codes = []
            answer_keys = {}
            
            if not df.empty:
                # Lấy cột đầu tiên cho exam codes
                first_col = df.iloc[:, 0].dropna()
                
                for index, value in enumerate(first_col):
                    if pd.notna(value):
                        str_value = str(value).replace('.0', '')
                        # Chỉ lấy những giá trị là số (mã đề)
                        if str_value.isdigit() or (str_value.replace('.', '').isdigit()):
                            exam_codes.append(str_value)
                            
                            # Lấy đáp án cho mã đề này (từ cột thứ 2 trở đi)
                            if index < len(df):
                                answers = []
                                # Lấy đáp án từ các cột từ 1 trở đi (bỏ qua cột 0 là exam code)
                                for col_idx in range(1, len(df.columns)):
                                    if index < len(df):
                                        answer_val = df.iloc[index, col_idx]
                                        if pd.notna(answer_val):
                                            answer_str = str(answer_val).strip().upper()
                                            # Chỉ lấy A, B, C, D
                                            if answer_str in ['A', 'B', 'C', 'D']:
                                                answers.append(answer_str)
                                            else:
                                                answers.append('')  # Empty answer for invalid values
                                        else:
                                            answers.append('')  # Empty answer for NaN
                                
                                answer_keys[str_value] = answers
                                logger.info(f"Exam code {str_value}: {len(answers)} answers")
                
                logger.info(f"Found exam codes: {exam_codes}")
                logger.info(f"Answer keys sample: {dict(list(answer_keys.items())[:2])}")  # Show first 2 for debugging
            
            return { 
                'message': 'Tải lên đáp án thành công!', 
                'filename': filename,
                'examCodes': exam_codes,
                'answerKeys': answer_keys
            }
                
        except Exception as excel_error:
            logger.error(f"Error reading Excel file: {excel_error}")
            return JSONResponse({'error': f'Lỗi đọc file Excel: {str(excel_error)}'}, status_code=500)
        
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        return JSONResponse({'error': f'Lỗi xử lý file: {str(e)}'}, status_code=500)

# Upload danh sách sinh viên (Excel)
@router.post('/api/upload_student_list')
async def upload_student_list(file: UploadFile = File(...)):
    logger.info(f"Received student list upload request: {file.filename}")
    
    if not file:
        logger.error("No file provided")
        return JSONResponse({'error': 'No file part'}, status_code=400)
    
    # Kiểm tra định dạng file
    if not file.filename.endswith('.xlsx'):
        logger.error(f"Invalid file format: {file.filename}")
        return JSONResponse({'error': 'Chỉ chấp nhận file Excel (.xlsx)'}, status_code=400)
    
    try:
        filename = file.filename
        save_path = os.path.join('uploads', 'student', filename)
        
        logger.info(f"Saving student list file to: {save_path}")
        
        # Đọc file content
        content = await file.read()
        logger.info(f"File size: {len(content)} bytes")
        
        # Đảm bảo thư mục tồn tại
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        
        # Lưu file
        with open(save_path, 'wb') as f:
            f.write(content)
        
        logger.info("Student list file saved successfully")
        
        # Xử lý file Excel sử dụng process_df_student
        try:
            result = process_df_student(save_path)
            
            if 'error' in result:
                logger.error(f"Error processing student list: {result['error']}")
                return JSONResponse({'error': result['error']}, status_code=400)
            
            # Lưu df_parts vào file tạm
            temp_dir = os.path.join('uploads', 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            df_parts_file = os.path.join(temp_dir, f'df_parts_{filename}.json')
            with open(df_parts_file, 'w', encoding='utf-8') as f:
                json.dump({key: part.to_dict() for key, part in result['df_parts'].items()}, f, ensure_ascii=False)
            
            logger.info(f"Saved df_parts to: {df_parts_file}")
            
            # Tạo danh sách students cho response
            students = []
            for part_key, part_df in result['df_parts'].items():
                for _, row in part_df.iterrows():
                    student = {
                        'stt': int(row['STT']),
                        'mssv': str(row['MSSV']),
                        'hoTen': str(row.get('Ten', '')) if 'Ten' in part_df.columns else ''
                    }
                    students.append(student)
            
            logger.info(f"Processed {len(students)} students in {len(result['df_parts'])} parts")
            
            return {
                'message': f'Tải danh sách thành công từ file {filename}',
                'filename': filename,
                'studentCount': len(students),
                'students': students,
                'df_parts_file': df_parts_file,
                'num_parts': len(result['df_parts']),
                'parts_info': {key: (part_df['STT'].iloc[0], part_df['STT'].iloc[-1]) for key, part_df in result['df_parts'].items()}
            }
            
        except Exception as excel_error:
            logger.error(f"Error processing Excel file: {excel_error}")
            return JSONResponse({'error': f'Lỗi xử lý file Excel: {str(excel_error)}'}, status_code=500)
        
    except Exception as e:
        logger.error(f"Error processing student list file: {str(e)}")
        return JSONResponse({'error': f'Lỗi xử lý file: {str(e)}'}, status_code=500)

# Upload ảnh bài làm
@router.post('/api/upload_exam_images')
async def upload_exam_images(images: List[UploadFile] = File(...)):
    logger.info(f"Received {len(images)} image uploads")
    
    saved_files = []
    for i, file in enumerate(images):
        if not file.filename:
            logger.warning(f"File {i} has no filename, skipping")
            continue
            
        filename = file.filename
        save_path = os.path.join('uploads', 'images', filename)
        
        # Đảm bảo thư mục tồn tại
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        
        try:
            content = await file.read()
            with open(save_path, 'wb') as f:
                f.write(content)
            saved_files.append(filename)
            logger.info(f"Saved image {i+1}/{len(images)}: {filename} ({len(content)} bytes)")
        except Exception as e:
            logger.error(f"Error saving file {filename}: {str(e)}")
            continue
    
    logger.info(f"Successfully uploaded {len(saved_files)}/{len(images)} images")
    return { 
        'message': f'Uploaded {len(saved_files)} images successfully', 
        'files': saved_files 
    }

# Xử lý ảnh và trả về kết quả thực tế
@router.post('/api/process_images')
async def process_images(request: Request):
    try:
        data = await request.json()
        answer_key_filename = data.get('answer_key_filename')
        student_list_filename = data.get('student_list_filename')
        image_filenames = data.get('image_filenames')
        room = data.get('room')
        
        logger.info(f"Processing images with params: answer_key={answer_key_filename}, student_list={student_list_filename}, images={len(image_filenames)}, room={room}")
        
        if not answer_key_filename or not student_list_filename or not image_filenames or not room:
            return JSONResponse({'error': 'Missing required parameters'}, status_code=400)
        
        # Đường dẫn file
        data_path_process = os.path.join('uploads', 'key', answer_key_filename)
        df_parts_file = os.path.join('uploads', 'student', student_list_filename)
        phongthi = room
        
        logger.info(f"Checking files - data_path_process: {data_path_process}")
        logger.info(f"df_parts_file: {df_parts_file}")
        logger.info(f"phongthi: {phongthi}")
        
        if not os.path.exists(data_path_process):
            return JSONResponse({'error': 'Không có file đáp án được tải lên hoặc xử lý.'}, status_code=400)
        if not os.path.exists(df_parts_file):
            return JSONResponse({'error': 'Không có dữ liệu danh sách học sinh.'}, status_code=400)
        
        # Đọc df_parts từ file JSON (được tạo từ upload_student_list)
        try:
            # Tìm file df_parts tương ứng với student_list_filename
            df_parts_file = os.path.join('uploads', 'temp', f'df_parts_{student_list_filename}.json')
            
            if not os.path.exists(df_parts_file):
                logger.error(f"df_parts file not found: {df_parts_file}")
                return JSONResponse({'error': 'File df_parts không tồn tại. Vui lòng upload lại danh sách sinh viên.'}, status_code=400)
            
            with open(df_parts_file, 'r', encoding='utf-8') as f:
                df_parts_data = json.load(f)
            
            # Chuyển đổi JSON thành DataFrame
            df_parts = {}
            for key, data in df_parts_data.items():
                df = pd.DataFrame(data)
                df = df.reset_index(drop=True)  # Reset index để đảm bảo index là số nguyên
                df_parts[key] = df
            
            logger.info(f"Loaded df_parts from JSON: {list(df_parts.keys())}")
            
            # Log chi tiết từng part
            for part_name, df in df_parts.items():
                logger.info(f"Part '{part_name}': shape={df.shape}, columns={list(df.columns)}, index_type={type(df.index[0]) if len(df) > 0 else 'empty'}")
                logger.info(f"Part '{part_name}' first few rows: {df.head(2).to_dict()}")
                
        except Exception as e:
            logger.error(f"Error reading df_parts file: {e}")
            return JSONResponse({'error': f'Lỗi đọc file df_parts: {str(e)}'}, status_code=400)
        
        # Xử lý phòng thi - tìm part phù hợp
        available_parts = list(df_parts.keys())
        logger.info(f"Available parts: {available_parts}")
        
        # Tìm part phù hợp với phòng thi
        part_key = None
        room_number = phongthi.replace('A', '').replace('B', '').replace('C', '')
        try:
            room_num = int(room_number)
            part_key = f"df_part{room_num}"
            if part_key not in available_parts:
                # Nếu không tìm thấy, sử dụng part đầu tiên
                part_key = available_parts[0]
                logger.info(f"Room {phongthi} not found, using first part: {part_key}")
        except ValueError:
            # Nếu không parse được số phòng, sử dụng part đầu tiên
            part_key = available_parts[0] if available_parts else None
            logger.info(f"Invalid room format, using first part: {part_key}")
        
        if part_key is None:
            return JSONResponse({'error': f'Không tìm thấy part phù hợp cho phòng thi {phongthi}.'}, status_code=400)
        
        logger.info(f"Using part: {part_key}")
        df_part = df_parts[part_key]
        logger.info(f"df_part columns: {df_part.columns.tolist()}")
        
        required_columns = ['STT', 'MSSV']
        missing_columns = [col for col in required_columns if col not in df_part.columns]
        if missing_columns:
            return JSONResponse({'error': f'File danh sách học sinh thiếu cột: {", ".join(missing_columns)}'}, status_code=400)
        
        student_ids = df_part['MSSV'].astype(str).tolist()
        if 'HoDem' in df_part.columns and 'Ten' in df_part.columns:
            student_names = (df_part['HoDem'].astype(str) + ' ' + df_part['Ten'].astype(str)).tolist()
        elif 'Ten' in df_part.columns:
            student_names = df_part['Ten'].astype(str).tolist()
        else:
            student_names = []
        stt_list = df_part['STT'].astype(str).tolist()
        
        logger.info(f"Processing for {phongthi}, student_ids: {student_ids}, student_names: {student_names}, stt_list: {stt_list}")
        
        # Đọc file đáp án
        try:
            df_key = pd.read_excel(data_path_process)
            logger.info(f"df_key shape: {df_key.shape}")
            logger.info(f"df_key indices: {list(df_key.index)}")
        except Exception as e:
            logger.error(f"Error reading answer key file: {e}")
            return JSONResponse({'error': 'Không thể xử lý file đáp án.'}, status_code=400)
        
        # Tính số câu hỏi (bỏ cột đầu tiên là mã đề)
        num_questions = len(df_key.columns) - 1
        logger.info(f"Số câu hỏi: {num_questions} (total columns: {len(df_key.columns)})")
        
        students = []
        successful_recognitions = 0
        
        for image_filename in image_filenames:
            try:
                image_path = os.path.join('uploads', 'images', image_filename)
                logger.info(f"Processing image: {image_path}")
                
                if not os.path.exists(image_path):
                    logger.warning(f"Image file not found: {image_path}")
                    continue
                
                # Xử lý ảnh và lấy tọa độ vùng phiếu thi
                processing_result = image_processing(image_path)
                paths = processing_result.get('paths', {})
                
                temp_file_name = os.path.basename(image_path).split('.')[0]
                temp_file_name = temp_file_name.replace('\t', '').replace('\\', '/')
                
                # Tạo thư mục temp nếu chưa có
                temp_dir = os.path.join('uploads', 'images', 'temp', temp_file_name)
                os.makedirs(temp_dir, exist_ok=True)
                
                # Detect các thông tin từ ảnh
                code_box_path = paths.get('code_box', os.path.join(temp_dir, 'code_box_bounding_box.jpg'))
                name_student_path = paths.get('name', os.path.join(temp_dir, 'name_bounding_box.jpg'))
                grading_path = paths.get('table_grading', os.path.join(temp_dir, 'table_grading_bounding_box.jpg'))
                id_student_path = paths.get('id_student', os.path.join(temp_dir, 'id_student.jpg'))
                index_student_path = paths.get('index_student', os.path.join(temp_dir, 'index_student.jpg'))
                
                logger.info(f"Processing paths: code_box={normalize_path(code_box_path)}, name={normalize_path(name_student_path)}, grading={normalize_path(grading_path)}")
                
                # Detect thông tin từ các vùng ảnh (raw detection)
                exam_code = detect_code_box(code_box_path)
                raw_name = detect_name_student(name_student_path, student_names)
                raw_id = detect_id_student(id_student_path, student_ids)
                raw_index = detect_index_student(index_student_path)
                
                # Xử lý ảnh để lấy đáp án bằng YOLO model với bounding boxes
                try:
                    logger.info(f"Starting YOLO processing for {grading_path}")
                    # Gọi predict_grade với save_processed_image=True để tạo ảnh với bounding boxes
                    processed_image_path, student_result = predict_grade(grading_path, save_processed_image=True)
                    logger.info(f"YOLO processing completed. Processed image: {processed_image_path}")
                    logger.info(f"Student result keys: {list(student_result.keys()) if student_result else 'None'}")
                except Exception as e:
                    logger.error(f"Error in predict_grade: {str(e)}")
                    raise Exception(f"Error in YOLO processing: {str(e)}")

                if not student_result:
                    logger.error("No answers detected by YOLO model")
                    raise Exception("No answers detected by YOLO model")
                
                # Lấy STT đầu tiên nếu có
                raw_stt = raw_index[0] if raw_index else None
                
                # Validate và correct thông tin sinh viên
                validation_result = validate_and_correct_student_info(
                    detected_name=raw_name,
                    detected_mssv=raw_id,
                    detected_stt=raw_stt,
                    df_students=df_part
                )
                
                # Lấy thông tin đã được correct
                corrected_name = validation_result['name']
                corrected_mssv = validation_result['mssv']
                corrected_stt = validation_result['stt']
                correction_status = validation_result['status']
                correction_reason = validation_result['correction_reason']
                
                # Kiểm tra kết quả chấm điểm
                if student_result is None:
                    logger.warning(f"Không thể xử lý bảng chấm điểm tại {grading_path}")
                    answers = [''] * num_questions
                else:
                    # Chuyển student_result thành danh sách câu trả lời
                    answers = [student_result.get(i, '') for i in range(1, num_questions + 1)]
                    
                    # Kiểm tra độ dài câu trả lời
                    if len(answers) != num_questions:
                        logger.warning(f"Số câu trả lời ({len(answers)}) không khớp với số câu hỏi ({num_questions})")
                        answers = answers + [''] * (num_questions - len(answers)) if len(answers) < num_questions else answers[:num_questions]
                
                # Tính điểm
                score = calculate_score(answers, df_key, exam_code)
                
                # Kiểm tra có vấn đề gì không
                has_issue = (
                    exam_code not in [str(idx).replace('.0', '') for idx in df_key.index] or 
                    not corrected_name or 
                    not corrected_mssv or 
                    not corrected_stt or
                    correction_status != 'exact_match'
                )
                
                # Tạo đường dẫn cho ảnh processed
                processed_filename = os.path.basename(processed_image_path)
                
                student = {
                    'id': corrected_mssv,
                    'name': corrected_name,
                    'testVariant': exam_code,
                    'score': score,
                    'answers': answers,
                    'image': normalize_path(image_filename),
                    'imageName': temp_file_name,  # Tên folder cho processed images
                    'processedGradingImage': normalize_path(f"temp/{temp_file_name}/{processed_filename}"),  # Đường dẫn ảnh đã xử lý với bounding boxes
                    'has_issue': has_issue,
                    'num_questions': num_questions,
                    'index_student': corrected_stt or 'N/A',
                    'correction_status': correction_status,
                    'correction_reason': correction_reason,
                    'raw_detection': {
                        'name': raw_name,
                        'mssv': raw_id,
                        'stt': raw_stt
                    }
                }
                
                students.append(student)
                if corrected_name and corrected_mssv and corrected_stt and score is not None and score > 0:
                    successful_recognitions += 1
                    
            except Exception as e:
                logger.error(f"Error processing {image_filename}: {e}")
                continue
        
        total_students = len(students)
        recognition_rate = (successful_recognitions / total_students * 100) if total_students > 0 else 0
        
        logger.info(f"Processing completed. Total students: {total_students}, Recognition rate: {recognition_rate:.2f}%")
        
        return {
            'message': 'Processing completed',
            'results': students,
            'total_students': total_students,
            'recognition_rate': round(recognition_rate, 2),
            'num_questions': num_questions
        }
        
    except Exception as e:
        logger.error(f"Error processing images: {str(e)}")
        return JSONResponse({'error': str(e)}, status_code=500)

# Tải file kết quả (Excel)
@router.get('/api/download_result/{filename}')
async def download_result(filename: str):
    directory = os.path.join('backend', 'output_steps')
    file_path = os.path.join(directory, filename)
    if not os.path.exists(file_path):
        return JSONResponse({'error': 'File not found'}, status_code=404)
    return FileResponse(file_path, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') 

# Lấy danh sách phòng thi
@router.get('/api/exam_rooms')
async def get_exam_rooms():
    try:
        # TODO: Lấy từ database hoặc config file
        # Hiện tại sử dụng danh sách cố định
        rooms = [
            {"id": "A101", "name": "Phòng A101", "capacity": 45},
            {"id": "A102", "name": "Phòng A102", "capacity": 45},
            {"id": "A103", "name": "Phòng A103", "capacity": 45},
            {"id": "B201", "name": "Phòng B201", "capacity": 45},
            {"id": "B202", "name": "Phòng B202", "capacity": 45},
            {"id": "C301", "name": "Phòng C301", "capacity": 45},
            {"id": "C302", "name": "Phòng C302", "capacity": 45},
        ]
        logger.info(f"Returning {len(rooms)} available exam rooms")
        return {"rooms": rooms}
    except Exception as e:
        logger.error(f"Error getting exam rooms: {str(e)}")
        return JSONResponse({'error': f'Lỗi lấy danh sách phòng thi: {str(e)}'}, status_code=500)

# Tính toán phòng thi dựa trên số lượng sinh viên
@router.post('/api/calculate_rooms')
async def calculate_rooms(request: Request):
    try:
        data = await request.json()
        student_count = data.get('student_count', 0)
        
        if student_count <= 0:
            return JSONResponse({'error': 'Số lượng sinh viên phải lớn hơn 0'}, status_code=400)
        
        # Logic chia phòng thi
        if student_count <= 45:
            # 1 phòng
            rooms_needed = 1
            students_per_room = [student_count]
        elif student_count <= 90:
            # 2 phòng
            rooms_needed = 2
            first_room = (student_count + 1) // 2  # Làm tròn lên
            second_room = student_count - first_room
            students_per_room = [first_room, second_room]
        else:
            # 3 phòng trở lên
            rooms_needed = 3
            base_per_room = student_count // 3
            remainder = student_count % 3
            students_per_room = [base_per_room + (1 if i < remainder else 0) for i in range(3)]
        
        # Tạo danh sách phòng thi
        available_rooms = [
            {"id": "Phong1", "name": "Phòng 1"},
            {"id": "Phong2", "name": "Phòng 2"},
            {"id": "Phong3", "name": "Phòng 3"}
        ]
        selected_rooms = []
        for i in range(rooms_needed):
            if i < len(available_rooms):
                room = available_rooms[i].copy()
                room['student_count'] = students_per_room[i]
                selected_rooms.append(room)
        
        return {
            'student_count': student_count,
            'rooms_needed': rooms_needed,
            'rooms': selected_rooms,
            'students_per_room': students_per_room
        }
        
    except Exception as e:
        logger.error(f"Error calculating rooms: {str(e)}")
        return JSONResponse({'error': f'Lỗi tính toán phòng thi: {str(e)}'}, status_code=500) 

# Serve uploaded images
@router.get('/api/images/{filename}')
async def get_image(filename: str):
    try:
        file_path = os.path.join('uploads', 'images', filename)
        if not os.path.exists(file_path):
            return JSONResponse({'error': 'Image not found'}, status_code=404)
        return FileResponse(file_path)
    except Exception as e:
        logger.error(f"Error serving image {filename}: {str(e)}")
        return JSONResponse({'error': 'Error serving image'}, status_code=500)

# Serve processed images (table grading results)
@router.get('/api/processed_images/{image_folder}/{filename}')
async def get_processed_image(image_folder: str, filename: str):
    try:
        file_path = os.path.join('uploads', 'images', 'temp', image_folder, filename)
        if not os.path.exists(file_path):
            return JSONResponse({'error': 'Processed image not found'}, status_code=404)
        return FileResponse(file_path)
    except Exception as e:
        logger.error(f"Error serving processed image {image_folder}/{filename}: {str(e)}")
        return JSONResponse({'error': 'Error serving processed image'}, status_code=500)

# Export results to original Excel file
@router.post('/api/export_to_original_excel')
async def export_to_original_excel(request: Request):
    try:
        data = await request.json()
        results = data.get('results', [])
        student_filename = data.get('student_filename', '')
        
        logger.info(f"Export to original Excel request - student_filename: {student_filename}, results count: {len(results)}")
        
        if not results or not student_filename:
            return JSONResponse({'error': 'Missing results or student filename'}, status_code=400)
        
        # Đường dẫn file Excel gốc
        original_file_path = os.path.join('uploads', 'student', student_filename)
        
        if not os.path.exists(original_file_path):
            logger.error(f"Original student file not found: {original_file_path}")
            return JSONResponse({'error': 'Original student file not found'}, status_code=400)
        
        try:
            # Sử dụng openpyxl để giữ nguyên format file Excel
            from openpyxl import load_workbook
            
            # Load workbook với tất cả formatting
            workbook = load_workbook(original_file_path)
            worksheet = workbook.active
            
            logger.info(f"Loaded Excel workbook: {original_file_path}")
            logger.info(f"Worksheet dimensions: {worksheet.max_row} rows x {worksheet.max_column} cols")
            
            # Tìm vị trí cột ThangDiem4 và vùng dữ liệu sinh viên
            # Dựa vào cấu trúc: skiprows=6, header=[0, 1, 2]
            header_row = 7  # Row 7 trong Excel (0-indexed + 6 + 1)
            data_start_row = 8  # Dữ liệu bắt đầu từ row 8
            
            # Mặc định cột MSSV là cột B (cột thứ 2) và ThangDiem4 là cột M (cột thứ 13)
            mssv_col = 2
            thangdiem4_col = 13
            logger.info(f"Using default columns - MSSV: B (column {mssv_col}), ThangDiem4: M (column {thangdiem4_col})")
            
            # Tạo dictionary để map MSSV -> Điểm (làm tròn đến 1 chữ số thập phân)
            score_map = {}
            for result in results:
                mssv = str(result.get('mssv', '')).strip()
                diem = result.get('diem', 0)
                if mssv and mssv != 'N/A':
                    # Làm tròn điểm đến 1 chữ số thập phân
                    score_map[mssv] = round(float(diem), 1)
            
            logger.info(f"Created score map for {len(score_map)} students")
            
            # Update điểm vào cột ThangDiem4 với giữ nguyên format
            updated_count = 0
            for row in range(data_start_row, worksheet.max_row + 1):
                mssv_cell = worksheet.cell(row=row, column=mssv_col)
                if mssv_cell.value:
                    try:
                        mssv_str = str(int(mssv_cell.value)) if isinstance(mssv_cell.value, (int, float)) else str(mssv_cell.value).strip()
                        if mssv_str in score_map:
                            thangdiem4_cell = worksheet.cell(row=row, column=thangdiem4_col)
                            # Ghi điểm đã được làm tròn đến 1 chữ số thập phân
                            thangdiem4_cell.value = score_map[mssv_str]
                            updated_count += 1
                            logger.debug(f"Updated MSSV {mssv_str}: {score_map[mssv_str]}")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Could not process MSSV at row {row}: {e}")
                        continue
            
            logger.info(f"Updated scores for {updated_count}/{len(score_map)} students")
            
            # Tạo file output mới với timestamp
            timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"ketqua_chamthi_{timestamp}.xlsx"
            output_dir = os.path.join('uploads', 'results')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, output_filename)
            
            # Lưu workbook với tất cả formatting gốc
            workbook.save(output_path)
            
            logger.info(f"Saved updated Excel file with original formatting: {output_path}")
            
            return {
                'message': f'Đã xuất kết quả vào file Excel gốc với đầy đủ format!',
                'filename': output_filename,
                'output_path': output_path,
                'updated_count': updated_count,
                'total_students': len(results),
                'mssv_column': mssv_col,
                'thangdiem4_column': thangdiem4_col
            }
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            return JSONResponse({'error': f'Error processing Excel file: {str(e)}'}, status_code=500)
        
    except Exception as e:
        logger.error(f"Error in export_to_original_excel: {str(e)}")
        return JSONResponse({'error': f'Error exporting to Excel: {str(e)}'}, status_code=500)

# Download result Excel file
@router.get('/api/download_result_excel/{filename}')
async def download_result_excel(filename: str):
    try:
        file_path = os.path.join('uploads', 'results', filename)
        if not os.path.exists(file_path):
            return JSONResponse({'error': 'Result file not found'}, status_code=404)
        return FileResponse(
            file_path, 
            filename=filename, 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error downloading result file {filename}: {str(e)}")
        return JSONResponse({'error': 'Error downloading file'}, status_code=500)